from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse

import openpyxl
from .models import Enquete
from .forms import EnqueteForm

#Accueil de l'application
from django.shortcuts import render
def index(request):
    return render(request, 'index.html')

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Enquete

@login_required(login_url='enquete:login')
def liste_enquetes(request):
    commune = request.GET.get('commune')
    activite = request.GET.get('activite')
    search = request.GET.get('search')

    enquetes = Enquete.objects.all().order_by('-created_at')

    # Appliquer les filtres
    if commune:
        enquetes = enquetes.filter(commune=commune)
    if activite:
        enquetes = enquetes.filter(activite_principale=activite)
    if search:
        enquetes = enquetes.filter(
            Q(nom_chef_atelier__icontains=search) |
            Q(prenom_chef_atelier__icontains=search) |
            Q(telephone_chef__icontains=search)
        )

    # Pagination
    paginator = Paginator(enquetes, 10)  # 10 éléments par page
    page = request.GET.get('page')
    enquetes_page = paginator.get_page(page)

    # Récupérer les valeurs pour les filtres (distinctes)
    communes = Enquete.objects.values_list('commune', flat=True).distinct()
    activites = Enquete.TYPE_ACTIVITE_CHOICES  # Si tu utilises choices dans ton modèle

    context = {
        'enquetes': enquetes_page,
        'communes': communes,
        'activites': activites,
        'is_paginated': enquetes_page.has_other_pages(),
        'page_obj': enquetes_page,
    }
    return render(request, 'liste.html', context)




# Vue pour ajouter une enquête
def ajouter_enquete(request):
    if request.method == 'POST':
        form = EnqueteForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('enquete:liste_enquetes')
    else:
        form = EnqueteForm()
    return render(request, 'formulaire.html', {'form': form, 'titre': 'Ajouter'})

# Vue pour voir les détails d'une enquête
def detail_enquete(request, pk):
    enquete = get_object_or_404(Enquete, pk=pk)
    return render(request, 'detail.html', {'enquete': enquete})

# Vue pour modifier une enquête
def modifier_enquete(request, pk):
    enquete = get_object_or_404(Enquete, pk=pk)
    if request.method == 'POST':
        form = EnqueteForm(request.POST, request.FILES, instance=enquete)
        if form.is_valid():
            form.save()
            return redirect('enquete:detail_enquete', pk=pk)
    else:
        form = EnqueteForm(instance=enquete)
    return render(request, 'formulaire.html', {'form': form, 'titre': 'Modifier'})

# Vue pour supprimer une enquête
def supprimer_enquete(request, pk):
    enquete = get_object_or_404(Enquete, pk=pk)
    if request.method == 'POST':
        enquete.delete()
        return redirect('enquete:liste_enquetes')
    return render(request, 'supprimer.html', {'enquete': enquete})


#connection et déconnexion des utilisateurs
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import View
from .forms import LoginForm

class LoginView(View):
    template_name = 'enquete/login.html'
    form_class = LoginForm

    def get(self, request):
        if request.user.is_authenticated:
            return redirect('enquete:liste_enquetes')
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                next_url = request.POST.get('next', 'enquete:liste_enquetes')
                return redirect(next_url)
            else:
                messages.error(request, "Nom d'utilisateur ou mot de passe incorrect.")
        return render(request, self.template_name, {'form': form})

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "Vous avez été déconnecté avec succès.")
    return redirect('login')



# EXPORTATION DES ENQUÊTES EN FORMAT EXCEL
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.contrib.auth.decorators import login_required
from .models import Enquete

@login_required
def export_enquetes_excel(request):
    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Enquêtes"

    # Entêtes de colonnes correspondant à votre template
    columns = [
        'ID',
        'Chef d\'atelier',
        'Téléphone',
        'Commune',
        'Activité principale',
        'Revenu mensuel',
        'Date création',
        'Enquêteur',
        'Département',
        'Quartier',
        'Sexe',
        'Date naissance',
        'CNI'
    ]
    ws.append(columns)

    # Appliquer les mêmes filtres que dans la liste
    enquetes = Enquete.objects.all()
    
    # Filtrage comme dans votre template
    commune = request.GET.get('commune')
    activite = request.GET.get('activite')
    search = request.GET.get('search')
    
    if commune:
        enquetes = enquetes.filter(commune=commune)
    if activite:
        enquetes = enquetes.filter(activite_principale=activite)
    if search:
        enquetes = enquetes.filter(
            models.Q(nom_chef_atelier__icontains=search) |
            models.Q(prenom_chef_atelier__icontains=search) |
            models.Q(telephone_chef__icontains=search)
        )

    # Remplissage des données
    for enquete in enquetes:
        ws.append([
            enquete.id,
            f"{enquete.nom_chef_atelier} {enquete.prenom_chef_atelier}",
            enquete.telephone_chef,
            enquete.commune,
            enquete.get_activite_principale_display(),
            enquete.get_revenu_mensuel_display(),
            enquete.created_at.strftime('%d/%m/%Y') if enquete.created_at else '',
            str(enquete.enqueteur) if enquete.enqueteur else '',
            enquete.departement,
            enquete.quartier,
            enquete.get_sexe_display(),
            enquete.date_naissance.strftime('%d/%m/%Y') if enquete.date_naissance else '',
            enquete.cni
        ])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=enquetes.xlsx'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response


# EXPORTATION DES ENQUÊTES Detaller EN FORMAT PDF
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def exporter_enquete_excel(request, pk):
    enquete = Enquete.objects.get(pk=pk)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="enquete_{enquete.id}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Enquête {enquete.id}"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2e6c80", end_color="2e6c80", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = ['Section', 'Champ', 'Valeur']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Fonction utilitaire pour ajouter des données
    def add_data(section, fields):
        for name, value in fields:
            if value not in [None, ""]:
                ws.append([section, name, str(value)])
    
    # 1. Informations Générales
    add_data('Informations Générales', [
        ('Date de l\'enquête', enquete.date_enquete.strftime('%d/%m/%Y') if enquete.date_enquete else ''),
        ('Enquêteur', str(enquete.enqueteur)),
        ('Département', enquete.departement),
        ('Commune', enquete.commune),
        ('Quartier', enquete.quartier),
        ('Coordonnées GPS', f"Latitude: {enquete.latitude}, Longitude: {enquete.longitude}"),
        ('Altitude', f"{enquete.altitude} mètres"),
        ('Précision', enquete.precision),
        ('Propriétaire', f"{enquete.nom_proprietaire} {enquete.prenom_proprietaire}"),
        ('Chef d\'atelier', f"{enquete.nom_chef_atelier} {enquete.prenom_chef_atelier}"),
        ('Téléphone du chef', enquete.telephone_chef)
    ])
    
    # 2. Démographie
    add_data('Démographie', [
        ('Sexe', enquete.get_sexe_display()),
        ('Date de naissance', enquete.date_naissance.strftime('%d/%m/%Y') if enquete.date_naissance else ''),
        ('Numéro CNI', enquete.cni or 'Non renseigné'),
        ('Niveau d\'étude', enquete.get_niveau_etude_display()),
        ('Autre niveau d\'étude', enquete.autre_niveau_etude if enquete.niveau_etude == 'autre' else 'N/A')
    ])
    
    # 3. Activité Professionnelle
    add_data('Activité Professionnelle', [
        ('Profession', enquete.profession),
        ('Activité principale', enquete.get_activite_principale_display()),
        ('Autre activité', enquete.autre_activite_principale if enquete.activite_principale == 'autre' else 'N/A'),
        ('Type d\'activité', enquete.get_type_activite_display()),
        ('Autre type d\'activité', enquete.autre_activite if enquete.type_activite == 'autre' else 'N/A'),
        ('Revenu journalier', enquete.get_revenu_journalier_display()),
        ('Revenu mensuel', enquete.get_revenu_mensuel_display())
    ])
    
    # 4. Statut d'Occupation
    add_data('Statut d\'Occupation', [
        ('Statut d\'occupation', enquete.get_statut_occupation_display()),
        ('Statut non propriétaire', enquete.get_statut_non_proprietaire_display() if enquete.statut_occupation == 'exploitant_non_proprietaire' else 'N/A'),
        ('Autre statut', enquete.autre_statut_non_proprietaire if hasattr(enquete, 'autre_statut_non_proprietaire') and enquete.statut_non_proprietaire == 'autre' else 'N/A'),
        ('Loyer mensuel', f"{enquete.loyer_mensuel} FCFA" if hasattr(enquete, 'loyer_mensuel') and enquete.statut_non_proprietaire == 'locataire' else 'N/A'),
        ('Année d\'installation', enquete.annee_installation),
        ('Mode d\'acquisition', enquete.get_acquisition_site_display()),
        ('Autre mode d\'acquisition', enquete.autre_acquisition if enquete.acquisition_site == 'autre' else 'N/A'),
        ('Déjà déguerpi', 'Oui' if enquete.deja_deguerpi else 'Non'),
        ('Lieu de provenance', enquete.lieu_provenance if enquete.deja_deguerpi else 'N/A'),
        ('Meilleur site', enquete.meilleur_site)
    ])
    
    # 5. Situation Juridique
    add_data('Situation Juridique', [
        ('Menaces d\'expulsion', 'Oui' if enquete.menaces_expulsion else 'Non'),
        ('Auteur des menaces', enquete.auteur_menaces if enquete.menaces_expulsion else 'N/A'),
        ('Nature juridique', enquete.get_nature_juridique_display()),
        ('Autre nature juridique', enquete.autre_nature_juridique if enquete.nature_juridique == 'autre' else 'N/A'),
        ('Superficie de l\'atelier', f"{enquete.superficie_atelier} m²"),
        ('Statut de l\'activité', enquete.get_statut_activite_display())
    ])
    
    # 6. Ressources Humaines
    add_data('Ressources Humaines', [
        ('Nombre total de personnes', enquete.nombre_total_personnes),
        ('Nombre de propriétaires', enquete.nombre_proprietaires),
        ('Nombre d\'ouvriers', enquete.nombre_ouvriers),
        ('Nombre d\'apprentis', enquete.nombre_apprentis),
        ('Nombre de stagiaires', enquete.nombre_stagiaires),
        ('Formation professionnelle', 'Oui' if enquete.formation_professionnelle else 'Non'),
        ('École de formation', enquete.ecole_formation if enquete.formation_professionnelle else 'N/A'),
        ('Formation des ouvriers', 'Oui' if enquete.formation_ouvriers else 'Non'),
        ('Nombre de personnes formées', enquete.nombre_personnes_formees if enquete.formation_ouvriers else 'N/A'),
        ('Durée d\'apprentissage', enquete.duree_apprentissage)
    ])
    
    # 7. Équipements
    add_data('Équipements', [
        ('Électricité', 'Oui' if enquete.electricite else 'Non'),
        ('Eau courante', 'Oui' if enquete.eau_courante else 'Non'),
        ('Espace de stockage', 'Oui' if enquete.espace_stockage else 'Non'),
        ('Matériel de levage', 'Oui' if enquete.materiel_levage else 'Non'),
        ('Banc de diagnostic', 'Oui' if enquete.banc_diagnostic else 'Non'),
        ('Outillage spécialisé', 'Oui' if enquete.outillage_specialise else 'Non'),
        ('Type d\'outillage', enquete.type_outillage if enquete.outillage_specialise else 'N/A'),
        ('Accès internet', 'Oui' if enquete.acces_internet else 'Non')
    ])
    
    # 8. Besoins
    add_data('Besoins', [
        ('Besoins en diagnostic', 'Oui' if enquete.besoin_diagnostic else 'Non'),
        ('Besoins en véhicules hybrides', 'Oui' if enquete.besoin_vehicules_hybrides else 'Non'),
        ('Besoins en gestion d\'atelier', 'Oui' if enquete.besoin_gestion_atelier else 'Non'),
        ('Besoins en relation client', 'Oui' if enquete.besoin_relation_client else 'Non'),
        ('Besoins en froid/climatisation', 'Oui' if enquete.besoin_froid_climatisation else 'Non'),
        ('Autres besoins', enquete.autres_besoins_formation if enquete.autres_besoins_formation else 'Aucun'),
        ('Disposition à formaliser', 'Oui' if enquete.dispose_formaliser else 'Non'),
        ('Type d\'accompagnement', enquete.type_accompagnement if enquete.dispose_formaliser else 'N/A'),
        ('Affilié à une OPA', 'Oui' if enquete.affilie_opa else 'Non'),
        ('Nom de l\'OPA', enquete.nom_opa if enquete.affilie_opa else 'N/A')
    ])
    
    # 9. Gestion des déchets
    add_data('Gestion des Déchets', [
        ('Génère des déchets', 'Oui' if enquete.genere_dechets else 'Non'),
        ('Méthode de gestion', enquete.get_gestion_dechets_display() if enquete.genere_dechets else 'N/A'),
        ('Autre méthode', enquete.autre_gestion_dechets if enquete.gestion_dechets == 'autre' and enquete.genere_dechets else 'N/A'),
        ('Impact environnemental', enquete.get_impact_environnement_display()),
        ('Avantages positifs', enquete.avantages_positifs if enquete.impact_environnement == 'positifs' else 'N/A'),
        ('Inconvénients négatifs', enquete.inconvenients_negatifs if enquete.impact_environnement == 'negatifs' else 'N/A')
    ])
    
    # 10. Relations et Mobilité
    add_data('Relations et Mobilité', [
        ('Relations avec occupants', enquete.relations_occupants),
        ('Problèmes avec autorités', 'Oui' if enquete.problemes_autorites else 'Non'),
        ('Autorité concernée', enquete.get_autorite_probleme_display() if enquete.problemes_autorites else 'N/A'),
        ('Autre autorité', enquete.autre_autorite if enquete.autorite_probleme == 'autre' and enquete.problemes_autorites else 'N/A'),
        ('Avantages pour la commune', enquete.avantages_commune),
        ('Inconvénients pour la commune', enquete.inconvenients_commune),
        ('Opinion sur réglementation', enquete.opinion_reglementation),
        ('Prêt à déménager', 'Oui' if enquete.pret_demenager else 'Non'),
        ('Prêt à exercer hors Dakar', 'Oui' if enquete.pret_exercer_hors_dakar else 'Non'),
        ('Raison du refus', enquete.raison_refus_dakar if not enquete.pret_exercer_hors_dakar else 'N/A'),
        ('Équipements souhaités', enquete.equipements_souhaites)
    ])
    
    # 11. Accessibilité
    add_data('Accessibilité', [
        ('Difficultés d\'accès', enquete.difficultes_acces),
        ('Lieu d\'origine', enquete.lieu_origine),
        ('Temps d\'accès', enquete.temps_acces),
        ('Lieu de résidence clientèle', enquete.lieu_residence_clientele),
        ('Horaires d\'occupation', enquete.horaires_occupation)
    ])
    
    # 12. Commentaires
    add_data('Commentaires', [
        ('Commentaire de l\'enquêteur', enquete.commentaire_enqueteur if enquete.commentaire_enqueteur else 'Aucun'),
        ('Commentaires supplémentaires', enquete.commentaires_supplementaires if enquete.commentaires_supplementaires else 'Aucun'),
        ('Suggestions', enquete.suggestions if enquete.suggestions else 'Aucune')
    ])
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    
    wb.save(response)
    return response